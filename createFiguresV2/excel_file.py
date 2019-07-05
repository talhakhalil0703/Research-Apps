import openpyxl
from data_structure import BrainSection
from manage_patient_data import extract_mm_from_middle
from manage_patient_data import get_average

def create_averages_in_excel(patient_array, max_mm, do_not_run_mm, work_book):
    slopes_per_mm = [None] * max_mm * 2
    dorsal = BrainSection()
    ventral = BrainSection()
    sheet = work_book.create_sheet('Patients Average Data')
    x = 1
    while x < max_mm:
        extract_mm_from_middle(patient_array, dorsal, ventral, x, do_not_run_mm)
        slopes_per_mm[max_mm - x] = dorsal.exponents.copy()
        slopes_per_mm[max_mm + x - 1] = ventral.exponents.copy()
        sheet['A' + str(x)] = 'Points in: ' + str(x) + ' mm: ' + str(len(dorsal.average_error))
        sheet['B' + str(x)] = 'Dorsal Slope Average For : ' + str(x) + ' mm'
        sheet['C' + str(x)] = get_average(dorsal.average_exponents)
        sheet['D' + str(x)] = 'Dorsal Offset Average For : ' + str(x) + ' mm'
        sheet['E' + str(x)] = get_average(dorsal.average_offset)
        sheet['F' + str(x)] = 'Ventral Slope Average For : ' + str(x) + ' mm'
        sheet['G' + str(x)] = get_average(ventral.average_exponents)
        sheet['H' + str(x)] = 'Ventral Offset Average For : ' + str(x) + ' mm'
        sheet['I' + str(x)] = get_average(ventral.average_offset)
        x += 1

def append_trajectory_excel(patient, trajectory_number, work_book, mm_file):
    sheet_name = str(patient.name) + ' Trajectory ' + str(trajectory_number)

    if sheet_name not in work_book.sheetnames:
        work_book.create_sheet(sheet_name)
        sheet['A1'] = 'Slope'
        sheet['B1'] = 'Offset'
        sheet['C1'] = 'Error'
        sheet['D1'] = 'R2'
        sheet['E1'] = 'Peak'
        sheet['F1'] = 'Area'
    else:
        work_book.active = work_book.sheetnames.index(sheet_name)

    sheet = work_book.active
    sheet_index = sheet.max_row + 1
