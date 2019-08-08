#Clinic Program for FootHills Talha Khalil
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import sys
import os
import glob
import csv
from openpyxl import load_workbook
import openpyxl
import pandas as pd
from pyexcel.cookbook import merge_all_to_a_book

window = tk.Tk()
window.title('Slice Excel Files Into Individual Paitents')
data_path = Entry(window, width = 45)
data_path.grid(row = 0, column = 0, columnspan = 2)
data_path.insert(END, '')

def slice_excel():
    data_path_string = data_path.get()

    if data_path_string[-4:] != '.xls' and data_path_string[-5:] != '.xlsx' and data_path_string[-4:] != '.csv': #Checking to see if its an excel file
        print('Wrong file type!')
        return
    if data_path_string[-4:] == '.csv':
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(data_path_string, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save('dontmodify.xlsx')
        #merge_all_to_a_book(glob.glob(data_path_string), "dontmodify.xlsx")
        df = pd.read_excel('dontmodify.xlsx')
        xls_index = data_path_string.index('.csv')
    else:
        df = pd.read_excel(data_path_string)

    path = data_path_string[:xls_index] + ' Individual' + '.xlsx'

    #Opening the workbook
    wb = openpyxl.Workbook()
    wb.save(path)
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    grouped = df.groupby('regis_phn')

    #Splicing the data into patients and transposing it so its easier to read
    for patient in grouped.groups:
        new_group = grouped.get_group(patient).T
        new_group.to_excel(writer, sheet_name= str(patient), header = False)
        current_sheet = writer.sheets[str(patient)]
        current_sheet.column_dimensions['A'].width = 21
        
        for letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            current_sheet.column_dimensions[letter].width = 18

    #Saving and removing empty sheet
    print('Done')
    writer.save()
    writer.close()

def browse_clicked():
    window.filename = filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("CSV files","*.csv"), ("all files","*.*"), ("Excel spreadsheets","*.xls"),("Excel spreadsheets MAC","*.xlsx")))
    data_path.delete(0, END)
    data_path.insert(END, window.filename)
def exit_clicked():
    sys.exit(0)
def run_clicked():
    slice_excel()

browse_button = Button(window, text = 'BROWSE', command = browse_clicked)
run_button = Button(window, text = 'RUN', command = run_clicked)

browse_button.grid(row = 1, column = 0)
run_button.grid(row = 1, column = 1)
window.mainloop()
